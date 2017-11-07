# -*- coding: utf-8 -*-
import os
import re
import sys
import json
import regex
from copy import copy
import datetime
from collections import namedtuple, defaultdict
import inspect
import functools
import glob
import numpy as np

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

from pandas import DataFrame
import pandas as pd
import pdb

from core.utils import is_buildin, replace_str
from core.openpyxl_ext import insert_rows

from core.agg_func import agg_functions
from core.post_func import common_funcs

#  pd.describe_option('display')

pd.set_option('display.expand_frame_repr', False)
pd.set_option('display.max.colwidth', 20)
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', 500)
pd.set_option('display.unicode.east_asian_width', True)



def rpartial(func, *args):
    return lambda *a: func(*(a + args))


def compose(*functions):
    def compose2(f, g):
        return lambda x: f(g(x))
    return functools.reduce(compose2, functions, lambda x: x)


def register_custom_func(name, func):
    global common_funcs
    common_funcs[name] = func


class GroupBy():
    def __init__(self, params):
        self.by = params['by']
        self.agg = params['agg']
        self.eval_func()

    def eval_func(self):
        for name, func_with_arg in self.agg.items():
            title, func_str = func_with_arg[:2]
            if len(func_with_arg) > 1:
                func_args = func_with_arg[2:]

            if not is_buildin(func_str):
                try:
                    agg_func = agg_functions[func_str]
                    if func_args:
                        agg_func = rpartial(agg_func, *func_args)
                except NameError as ex:
                    raise
            else:
                agg_func = eval(func_str)
            self.agg[name] = [title, agg_func]

    def parse(self, df):
        group_by = self.by
        agg_dict = self.agg
        D = df.groupby(group_by, sort=False)
         
        dataframes = []
        for name, values in agg_dict.items():
            col, func = values
            df_grouped = pd.DataFrame(D[col].apply(lambda x: func(x)))
            df_grouped.columns = [name]
            dataframes.append(df_grouped)
        output = pd.concat(dataframes, axis=1)

        if len(group_by) > 1:
            output.index = output.index.droplevel(output.index.names[1:])

        return output


class Trim():
    def __init__(self, params):
        self.columns = params['columns']
        self.reset_index = params['reset_index']

    def parse(self, df):
        output = df.reset_index(df.index.name)[self.columns]
        output.index = range(1, len(output)+1)
        return output


class Extension():
    def __init__(self, params):
        self.title = params['title']
        self.output_type = params['type']
        self.input = params['input']
        self.func = params['func']

    def parse(self, df):
        input, func_name = self.input, self.func

        if self.title in df.columns:
            df_output = df.drop(self.title, axis=1)
        else:
            df_output = df

        output = []
        for index, row in df.iterrows():
            try:
                titles = df.columns.tolist()
                input_dict = dict(zip(titles, row))

                if func_name.startswith('template@'):

                    input_dict.keys()
                    template_string = func_name[9:]
                    p = regex.compile('{.*?\L<options>.*?}', options=input_dict.keys())
                    grouped = [m for m in p.finditer(template_string)]
                    spans = [list(g.span()) for g in grouped]

                    if not grouped:
                        each_output = template_string
                    else:
                        for key, value in input_dict.items():
                            locals()[key] = value

                        for key, value in common_funcs.items():
                            locals()[key] = value

                        values = []
                        for group in grouped:
                            span = group.span()
                            matched = group.group()
                            value = eval(matched[1:-1])
                            values.append(value)
                        each_output = replace_str(template_string, spans, values)

                        if self.output_type:
                            cast = {'int': int, 'float': float, 'list': json.loads}
                            each_output = cast[self.output_type](each_output)

                else:
                    func = global_func[func_name]
                    sig = inspect.signature(func)
                    arg_names = list(sig.parameters)
                    if len(arg_names)==1:
                        inputs = {arg_names[0]: input_dict[input[0]]}
                    else:
                        inputs = {k:v for k,v in input_dict.items() if k in arg_names}
                    each_output = func(**inputs)

                output.append(each_output)

            except AssertionError as ex:
                print('template後的參數未定義')
                import pdb; pdb.set_trace()
                #  raise AssertionError

            #  except Exception as ex:
                #  print('[Extension Parse Error]')
                #  raise ex


        df_ext = pd.DataFrame({self.title: output}, index = df.index)
        df_output = pd.concat((df_output, df_ext), axis=1)
        return df_output


class xlProcessor():
    def __init__(self, yml_file):
        self.yml = yml_file
        self.process()

    def process(self):
        import yaml
        content = defaultdict(str, yaml.load(open(self.yml, 'r').read()))
        self.converter = content['xlconverter']
        self.renderer = content['xlrenderer']
        
        if 'xlparser' in content:
            output = content['xlparser']
            outputs = defaultdict(list)
            for ot in output:
                output_name = ot['job_name'] 
                for each_op in ot['chains']:
                    op_args = defaultdict(str, each_op['args'])
                    item = op[each_op['type']](op_args)
                    outputs[output_name].append(item)
        else:
            outputs = None
        self.jobs = outputs
        return outputs


class xlConverter():
    def __init__(self, params):
        self.first_row, self.idx_colname = params['first_row'], params['idx_colname']

    def load_excel(self, filepath, sheet_name=0, resetindex=True):
        col_indexes, col_names = list(zip(*list(self.idx_colname.items())))        
        parse_cols = [c-1 for c in col_indexes]
        df = pd.read_excel(filepath, sheet_name=sheet_name,
                skiprows=self.first_row-1, header=None, usecols=parse_cols, names=col_names)
        df = df.reset_index(drop=True) if resetindex else df
        return df


class xlMerger(xlConverter):
    def __init__(self, params, source_path):
        xlConverter.__init__(self, params)
        self.source_path = source_path

    def merge_all(self, sheet_name=0, filename_excludes=None):
        files = glob.glob('{}/*xlsx*'.format(self.source_path))
        file_names = [os.path.basename(f) for f in files]
        if filename_excludes:
            file_names = list(filter(lambda f: f not in filename_excludes, file_names))
        else:
            file_names = file_names

        df_all = []
        for f in file_names:
            file_path = os.path.join(self.source_path, f)
            df_all.append(self.load_excel(file_path))
        df = pd.concat(df_all, axis=0)
        df.index = range(len(df))
        return df, file_names


class xlParserNew():
    def __init__(self, df, jobs):
        self.df = df
        self.jobs = jobs

    def parse(self):
        outputs = defaultdict(list)
        for t, operations in self.jobs.items():
            df = self.df
            for i, each_op in enumerate(operations):
                df = each_op.parse(df)
                outputs[t].append(df)
        self.outputs = outputs

    def show_process(self, jobs):
        for i, o in enumerate(jobs, 1):
            print('\n\noperation %d' % i)
            print(o.head())


class xlParser():
    def __init__(self, filepath, xl_template=None):
        self.template = xl_template
        self.filepath = filepath
        self.load_excel()
        self.parse()

    def parse(self):
        outputs = defaultdict(list)
        for t, operations in self.template.outputs.items():
            df = self.df
            for i, each_op in enumerate(operations):
                df = each_op.parse(df)
                outputs[t].append(df)
        self.outputs = outputs

    def show_process(self, outputs):
        for i, o in enumerate(outputs, 1):
            print('\n\noperation %d' % i)
            print(o.head())

    def load_excel(self, sheet_name=0, resetindex=True):
        col_indexes, title_names = list(zip(*list(self.template.title_dict.items())))        
        parse_cols = [c-1 for c in col_indexes]
        df = pd.read_excel(self.filepath, sheet_name=sheet_name,
                skiprows=self.template.first_row-1,
                header=None, usecols=parse_cols, names=title_names)
        df_final = df.reset_index(drop=True) if resetindex else df
        self.df = df_final


class xlRenderer():
    def __init__(self, params):
        self.first_row, self.idx_colname = params['first_row'], params['idx_colname']

    @classmethod
    def to_excel(cls, df, path):
        writer = pd.ExcelWriter(path, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        datetime_format = workbook.add_format({'num_format': "yyyy-mm-dd"})
        time_format = workbook.add_format({'num_format': "hh:mm"})

        for index, values in df.iterrows():
            for i, v in enumerate(values):
                if type(v) == pd.Timestamp:
                    worksheet.write_datetime('%s%s' % (chr(65+i), index + 2), v, datetime_format)
                elif type(v) == datetime.time:
                    cell = '%s%s' % (chr(65+i), index+2)
                    worksheet.write_datetime('%s%s' % (chr(65+i), index + 2), v, time_format)
        writer.save()

    def render_excel(self, df, saved_name, template_name):

        xls_file, xlsx_file = ['template/%s.%s' % (template_name, ext) for ext in ['xls', 'xlsx']]
        template = load_workbook(xlsx_file)

        rows = dataframe_to_rows(df, index=True, header=False)
        ws = template.active
        ws.insert_rows = insert_rows

        ws.insert_rows(ws, self.first_row, len(df))
        xls_from = xlrd.open_workbook(xls_file, formatting_info=True)
        xls_sheet = xls_from.sheet_by_index(0)
        color_index = lambda row, col: (
            xls_from.xf_list[xls_sheet.cell_xf_index(row,col)].background.pattern_colour_index)
        rgb = lambda row, col: xls_from.colour_map[color_index(row,col)]
        to_hexcode = lambda rgb_code: '00%s' % ''.join([('%x'% e).upper() for e in rgb_code])

        get_cell = lambda sht, row, col: sht['%s%d' % (get_column_letter(col), row)]

        #  for r_idx, row in enumerate(rows, self.first_row):
            #  for c_idx, value in enumerate(row, 1):
                #  get_cell(ws, r_idx, c_idx).value = value

        for r_idx, row in enumerate(rows, self.first_row):
            #  for c_idx, value in enumerate(row, 1):
            #  for c_idx in range(max(c_idxes)):
            for c_idx, df_colname in self.idx_colname.items():
                col_idx = df.columns.get_loc(df_colname)
                #  print(r_idx, c_idx, df_colname, col_idx)
                #  get_cell(ws, r_idx, c_idx).value = value
                get_cell(ws, r_idx, c_idx).value = row[col_idx+1]

        for row in ws.rows:
            max_contents_len = []
            for cell in row:
                r_idx, c_idx = cell.row, cell.col_idx

                max_content_len = len(str(get_cell(ws, r_idx, c_idx)))
                max_contents_len.append(max_content_len)

                #  if cell.has_style:
                new_cell = get_cell(ws, r_idx, c_idx)

                if r_idx >= self.first_row:
                    font = get_cell(ws, self.first_row, c_idx).font
                else:
                    font = cell.font
                new_cell.font = copy(font)
                new_cell.border = copy(cell.border)

                # uses xlrd to get cell styles
                if r_idx < self.first_row:
                    color = to_hexcode(rgb(r_idx-1, c_idx-1))
                else:
                    color = to_hexcode(rgb(self.first_row-1, c_idx-1))
                fill = PatternFill(fill_type='solid',start_color=color)
                new_cell.fill = copy(fill)

                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        template.save(saved_name)
        print('{} saved'.format(saved_name))


class XlManager(object):

    def __init__(self, yml, home=None):
        home = os.getcwd() if not home else home
        yml_path = '%s/%s' % (home, yml)
        self.process = xlProcessor(yml_file=yml_path)
        self.yml, self.home = yml_path, home
        self.df = None
        self.renderer = xlRenderer(self.process.renderer)
        self.register_func()
        
    def register_func(self):
        import func
        print('reigister_func')
        funcs = [f for f in dir(func) if f.startswith('exp')]
        for func_name in funcs:
            func_name_in_yml = func_name[4:]
            print('func {} registed.'.format(func_name))
            print('func name in yml: {}'.format(func_name_in_yml))
            register_custom_func(func_name_in_yml, getattr(func, func_name))

    def dummy(self):
        print(self.yml, self.home)

    def merge_all(self, sheet_name=0, filename_excludes=None, save_raw_excel=False):
        print(sheet_name, filename_excludes, save_raw_excel)

        source_path = os.path.join(self.home, 'source')
        self.merger = xlMerger(self.process.converter, source_path)
        df_source, files = self.merger.merge_all(sheet_name, filename_excludes)
        print('files merged: %s' % ','.join(files))
        if save_raw_excel:
            xlRenderer.to_excel(df_source, 'output/merged.xlsx')
        self.df = df_source
        return self

    def to_excel(self, saved_path):
        if self.df is not None:
            self.renderer.to_excel(self.df, saved_path)

    def render_excel(self, excel_template, saved_path):
        if self.df is not None:
            self.renderer.render_excel(self.df, saved_path, excel_template)

    def __str__(self):
        return str(self.df)

    def parse(self, job_name):
        if self.df is not None:
            xp = xlParserNew(self.df, self.process.jobs)
            xp.parse()
            job_output = xp.outputs[job_name]
            df_result = job_output[-1]
            self.df = df_result
        return self


op = {
    'group_by': GroupBy,
    'extend': Extension,
    'trim': Trim,
}

