# -*- coding: utf-8 -*-

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import os
import yaml
import json
import __main__
from copy import copy
import datetime
from collections import defaultdict
import functools

import xlrd
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from pretty_html_table import build_table


from explay.utils import pd_set_option
from explay.openpyxl_ext import insert_rows
from explay.parser import xlParser
from explay.merger import xlMerger, xlConverter


def compose(*functions):
    def compose2(f, g):
        return lambda x: f(g(x))

    return functools.reduce(compose2, functions, lambda x: x)


class xlRenderer:
    def __init__(self, home, params):
        self.home = home
        self.params = params
        self.first_row, self.idx_colname = params["first_row"], params["idx_colname"]

    def __repr__(self):
        msg = "[Renderer]\n"
        msg += yaml.dump(
            self.params, indent=True, allow_unicode=True, default_flow_style=False
        )
        return msg

    @classmethod
    def to_excel(cls, df, path):
        writer = pd.ExcelWriter(path, engine="xlsxwriter")
        df.to_excel(writer, sheet_name="Sheet1", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        datetime_format = workbook.add_format({"num_format": "yyyy-mm-dd"})
        time_format = workbook.add_format({"num_format": "hh:mm"})

        for index, values in df.iterrows():
            for i, v in enumerate(values):
                if type(v) == pd.Timestamp:
                    worksheet.write_datetime(
                        "%s%s" % (chr(65 + i), index + 2), v, datetime_format
                    )
                elif type(v) == datetime.time:
                    cell = "%s%s" % (chr(65 + i), index + 2)
                    worksheet.write_datetime(
                        "%s%s" % (chr(65 + i), index + 2), v, time_format
                    )
        writer.save()

    def render_excel(self, df, saved_name, template_name, template_dir="template"):

        #  xls_file, xlsx_file = ['template/%s.%s' % (template_name, ext) for ext in ['xls', 'xlsx']]
        xls_file, xlsx_file = [
            "%s/%s.%s" % (template_dir, template_name, ext) for ext in ["xls", "xlsx"]
        ]
        xlsx_file = os.path.join(self.home, xlsx_file)
        xls_file = os.path.join(self.home, xls_file)

        template = load_workbook(xlsx_file)

        #  rows = dataframe_to_rows(df, index=True, header=False)
        rows = dataframe_to_rows(df, index=False, header=False)

        ws = template.active
        ws.insert_rows = insert_rows
        ws.insert_rows(ws, self.first_row, len(df))

        xls_from = xlrd.open_workbook(xls_file, formatting_info=True)
        xls_sheet = xls_from.sheet_by_index(0)
        color_index = lambda row, col: (
            xls_from.xf_list[
                xls_sheet.cell_xf_index(row, col)
            ].background.pattern_colour_index
        )
        rgb = lambda row, col: xls_from.colour_map[color_index(row, col)]
        to_hexcode = lambda rgb_code: "00%s" % "".join(
            [("%x" % e).upper() for e in rgb_code]
        )

        get_cell = lambda sht, row, col: sht["%s%d" % (get_column_letter(col), row)]

        for r_idx, row in enumerate(rows, self.first_row):
            for c_idx, df_colname in self.idx_colname.items():
                col_idx = df.columns.get_loc(df_colname)
                #  get_cell(ws, r_idx, c_idx).value = row[col_idx + 1]
                get_cell(ws, r_idx, c_idx).value = row[col_idx]

        for row in ws.rows:
            max_contents_len = []
            for cell in row:
                if not hasattr(cell, "col_idx"):
                    continue
                r_idx, c_idx = cell.row, cell.col_idx

                max_content_len = len(str(get_cell(ws, r_idx, c_idx)))
                max_contents_len.append(max_content_len)

                #  if cell.has_style:
                new_cell = get_cell(ws, r_idx, c_idx)

                if r_idx >= self.first_row:
                    font = get_cell(ws, self.first_row, c_idx).font
                else:
                    font = cell.font
                new_cell.font = copy(font)
                new_cell.border = copy(cell.border)

                # uses xlrd to get cell styles
                if r_idx < self.first_row:
                    color = to_hexcode(rgb(r_idx - 1, c_idx - 1))
                else:
                    color = to_hexcode(rgb(self.first_row - 1, c_idx - 1))
                fill = PatternFill(fill_type="solid", start_color=color)
                new_cell.fill = copy(fill)

                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        saved_name = os.path.join(self.home, saved_name)
        if not os.path.isdir(saved_name):
            os.makedirs(os.path.dirname(saved_name), exist_ok=True)
        template.save(saved_name)
        print("{} saved".format(saved_name))


class xlTemplate:
    def __init__(self, params):
        self.params = params
        self._template = params["template"]
        self.output = params["output"]

    def __getitem__(self, index):
        return self._template[index]

    def __len__(self):
        return len(self._template)


class ExPlay:
    def __init__(self, home=None, proj_name=None):
        self.home = home or os.getcwd()
        self._parse_yml(proj_name)
        self.inputs = self._df_inputs()
        pd_set_option(max_colwidth=40, max_columns=15)

    def _parse_yml(self, proj_name):
        yml_file = os.path.join(self.home, f"{proj_name}.yml")

        #  each = yaml.load(open(yml_file, "r"), yaml.Loader)
        each = yaml.load(open(yml_file, "rb"), yaml.Loader)
        self._proj = each.get("xlproject", None)
        self._conv = each.get("xlconverter", None)
        self._merg = each.get("xlmerger", None)
        self._pars = each.get("xlparser", None)
        self._rend = each.get("xlrenderer", None)
        self._out = each.get("xloutput", None)

        # initialize converters
        self.converters = {}
        for k, v in self._conv.items():
            v.update({"name": k})
            self.converters[k] = xlConverter(v)

        # initialize parsers
        self.parsers = {}
        if self._pars:
            for name, param in self._pars.items():
                parser_init = defaultdict(str)
                self.parsers[name] = xlParser(name, param)

        #  self._renderer = xlRenderer(self._rend) if self._rend else None
        self._renderer = xlRenderer(self.home, self._rend) if self._rend else None
        self._template = xlTemplate(self._out) if self._out else None
        if self._proj:
            self._project = yaml.dump(self._proj, indent=True, default_flow_style=False)
        else:
            self._project = None

    def show_config(self):
        if self._project:
            print("************************")
            print("*       project        *")
            print("************************")
            print(self._project)

        if self._converter:
            print("************************")
            print("*       converter      *")
            print("************************")
            print(self._converter)

        if self._merg:
            print("************************")
            print("*         merger       *")
            print("************************")
            merg_print = yaml.dump(self._merg, allow_unicode=True, indent=True)
            print(merg_print)

        if self.parsers:
            print("************************")
            print("*         parser       *")
            print("************************")
            for each_parser in self.parsers:
                print(each_parser, "\n")

        if self._renderer:
            print("************************")
            print("*       renderer       *")
            print("************************")
            print(self._renderer)

    def _get_abs_source_path(self, xlsx_dir=None):
        if xlsx_dir:
            if os.path.isabs(xlsx_dir):
                source_path = xlsx_dir
            else:
                source_path = os.path.join(self.home, xlsx_dir)
        else:
            source_path = self.home
        return source_path

    def merge_sheets(self, conv_name, xlsx_path, sheet_names, save=False):
        print("sheets of file %s merged." % xlsx_path)
        xlsx_dir = self._get_abs_source_path(xlsx_path)
        #  self.merger = xlMerger(self._conv, xlsx_dir)
        #  df_merged = self.merger.merge_sheets(conv_name, xlsx_dir, sheet_names)
        merger = xlMerger(self._conv, xlsx_dir)
        df_merged = merger.merge_sheets(conv_name, xlsx_dir, sheet_names)
        if save:
            saved_path = "{}/{}_merged.xlsx".format(self.home, conv_name)
            xlRenderer.to_excel(df_merged, saved_path)
        #  return df_merged
        return merger

    def merge_files(self, conv_name, locations, sheet_name=0, save=False):
        print("merge_files")
        absfilepaths = []
        for each in locations:
            if os.path.isabs(each):
                absfilepaths.append(each)
            else:
                absfilepaths.append(os.path.join(os.path.abspath(self.home), each))
        #  self.merger = xlMerger(self._conv, source_path=self.home)
        #  df_merged = self.merger.merge_files(conv_name, absfilepaths, sheet_name)
        merger = xlMerger(self._conv[conv_name], source_path=self.home)
        df_merged = merger.merge_files(conv_name, absfilepaths, sheet_name)
        print("files merged")
        for e in absfilepaths:
            print(e)
        if save:
            saved_path = "{}/{}_merged.xlsx".format(self.home, conv_name)
            xlRenderer.to_excel(df_merged, saved_path)
        #  return df_merged
        return merger

    def merge_all(
        self, conv_name, xlsx_dir=None, sheet_name=0, excludes=None, save=False
    ):
        source_path = self._get_abs_source_path(xlsx_dir)
        #  self.merger = xlMerger(self._conv, source_path)
        #  df_merged, file_names = self.merger.merge_all(conv_name, sheet_name, excludes)

        merger = xlMerger(self._conv, source_path)
        df_merged, file_names = merger.merge_all(conv_name, sheet_name, excludes)
        print("files merged: %s" % ",".join(file_names))
        if save:
            saved_path = "{}/{}_merged.xlsx".format(self.home, conv_name)
            xlRenderer.to_excel(df_merged, saved_path)
        return df_merged
        #  return merger

    def _df_inputs(self):
        assert self._merg
        merged = {}
        self.mergers = {}
        for merger_name, each_merger in self._merg.items():
            each_merger = defaultdict(str, each_merger)

            # shared
            output, merge_type, conv_name, sheet_name = (
                each_merger["output"],
                each_merger["type"],
                each_merger["converter_name"],
                each_merger["sheet_name"],
            )
            sheet_name = sheet_name or 0  # default is the 1st sheet
            conv_params = self._conv[conv_name]
            merg_params = self._merg[merger_name]

            if merge_type == "merge_files":
                locations = each_merger["locations"]  # relative or abs
                absfilepaths = []
                for each in locations:
                    if os.path.isabs(each):
                        absfilepaths.append(each)
                    else:
                        absfilepaths.append(
                            os.path.join(os.path.abspath(self.home), each)
                        )
                merger = xlMerger(
                    merger_name, conv_params, merg_params, source_path=self.home
                )
                df_merged = merger.merge_files(conv_name, absfilepaths, sheet_name)

            elif merge_type == "merge_sheets":
                xlsx_path = each_merger["xlsx_path"]
                xlsx_dir = self._get_abs_source_path(xlsx_path)
                sheet_names = each_merger["sheet_names"]
                merger = xlMerger(
                    merger_name, conv_params, merg_params, source_path=xlsx_dir
                )
                df_merged = merger.merge_sheets(conv_name, xlsx_dir, sheet_names)

            elif merge_type == "merge_all":
                xlsx_dir = each_merger["xlsx_dir"]
                excludes = each_merger["excludes"]
                source_path = self._get_abs_source_path(xlsx_dir)
                merger = xlMerger(merger_name, conv_params, merg_params, source_path)
                df_merged = merger.merge_all(conv_name, sheet_name, excludes)

            merger.output = df_merged
            self.mergers[merger_name] = merger
            merged[output] = df_merged
        return merged

    def _run(self, each_proj):
        input_name = each_proj["input"]
        parser = self.parsers[each_proj["parser"]]
        df_input = self.inputs[input_name]
        result = parser(df_input)
        return result

    def run_proj(self, to_excel=True):
        components = [self.converters, self._merg, self.parsers, self._project]
        if not all(components):
            print("please define all explay components!")
            return
        self.results = {}
        for proj_name, each_proj in self._proj.items():
            print("==== each_proj ====", proj_name)
            self.results[proj_name] = self._run(each_proj)

        if to_excel:
            if self._out and self._renderer and self._template:
                self._render_excel()
            else:
                self._to_excel()
        else:
            for name, result in self.results.items():
                print("\nproj result: {} (first 10 rows)".format(name))
                print(result.head(10))

    def _to_excel(self):
        if not self._renderer:
            print("no renderer found.")
            return
        for proj_name, each_result in self.results.items():
            self._renderer.to_excel(each_result, "out_{}.xlsx".format(proj_name))

    def _render_excel(self):
        for e, e2 in zip(self._template.output, self._template.params["template"]):
            template_name = e["template"]
            template_dir = e2["dir"]
            proj_result = self.results[e["proj_result"]]
            path = e["path"]
            self._renderer.render_excel(proj_result, path, template_name, template_dir)

    def export_merged(self):
        out_xlsx = "merged.xlsx"
        writer = pd.ExcelWriter(out_xlsx, engine="xlsxwriter")
        for name, merger in self.mergers.items():
            merger.output.to_excel(writer, index=None, sheet_name=name)
        writer.save()

    def export_inputs(self):
        inputs = self._df_inputs()
        for input_name, each_df in inputs.items():
            setattr(__main__, input_name, each_df)

    def export_parsers(self):
        for k, each_parser in self.parsers.items():
            setattr(__main__, each_parser.name, each_parser)

    def export_html(self, projname, html_name="out", show_rows_max=10):
        proj = self._proj[projname]
        input = self.inputs[proj["input"]]
        parser = self.parsers[proj["parser"]]
        with open(f"{html_name}.html", "w") as f:
            title = f"<h4>Merged input</h4>"
            html = title + build_table(input, "blue_light", font_size="10px")
            f.write(html)
            for each, df in zip(parser, parser.output):
                print(each)
                details = json.dumps(dict(each.params["args"]), ensure_ascii=False)
                each_to_show = df[:show_rows_max]
                title = f"<h4>{each}</h4>"
                html = (
                    title
                    + details
                    + build_table(each_to_show, "blue_light", font_size="10px")
                )
                f.write(html)
